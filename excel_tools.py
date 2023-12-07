from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

path = r"C:\Users\dacan\OneDrive\Desktop\Projects\BARRAS\imagenes\DB.xlsx"

#Funciones: Status, Guardar

def status(path): #OK
    wb = load_workbook(path,read_only=True,data_only=True)
    ws = wb['DB']
    #Límite 5000
    limit = slice("A1","A5000")
    quantity = len([row[0].value for row in ws[limit]])
    wb.close()
    return quantity

def save2row(path,quantity,CCI,owner,name,colour,lengths,extra_info):
    wb = load_workbook(path)
    ws = wb['DB']

    #Writing:
    id = quantity+1
    ws.cell(row=id,column=1).value = id-1
    ws.cell(row=id,column=2).value = CCI
    ws.cell(row=id,column=3).value = owner
    ws.cell(row=id,column=4).value = name
    ws.cell(row=id,column=5).value = colour
    ws.cell(row=id,column=6).value = lengths
    ws.cell(row=id,column=7).value = extra_info

    border_style = Side(style='thin', color='000000')
    border = Border(top=border_style,bottom=border_style,left=border_style,right=border_style)

    for col in range(1,8):
        cell = ws.cell(row=id,column=col)
        cell.border = border

        if col <= 3:
            cell.alignment = Alignment(horizontal='center',vertical='center')

    wb.save(path)
    wb.close()

    return print("Datos enviados")

def find(path,barcode) -> None:
    wb = load_workbook(path)
    ws = wb['DB']
    limit = slice("B1","B20")
    cci_list = [row[0].value for row in ws[limit]]
    for i,cci in enumerate(cci_list):
        print(cci)
        if cci == int(barcode):
            print("############# RECURSO YA ASIGNADO #############")
            print(f"""ID: {i+1}
CCI: {ws.cell(i+1,2).value}
Asignado a: {ws.cell(i+1,3).value}
Color: {ws.cell(i+1,4).value}
Medidas: {ws.cell(i+1,5).value}
Extra: {ws.cell(i+1,6).value}""")
            break

def change(path,barcode):
    pass